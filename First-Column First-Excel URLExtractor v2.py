#!/usr/bin/python3
# Program cleans up XLSX files with URL and outputs in separate file
# File is placed in outputdirectory, program creates directory if it doesn't already exist
# This version does not use TKinter.

import os
from openpyxl import *
from openpyxl.utils.exceptions import InvalidFileException  # handles situation when no file is selected
import os

f1 = ""

class Compare():
    def __init__(self):
        self.file1 = ""
        # self.file2 = ""

    def setFile1(self, file1):
        print("File 1 set")
        self.file1 = file1
        return self.file1

    def process(self, f1filename):
        # Opening up the files
        file1 = load_workbook(self.file1)
        sheet1 = file1.active

        # Initialize sets for storing the (first name, last name) for the two files
        set1 = set()
        set2 = set()

        # Iterate through rows of each file and add to set
        for i in range(1, sheet1.max_row + 1):
            fname = sheet1.cell(row=i, column=1).value.lower()
            # lname = sheet1.cell(row=i, column=2).value.lower()
            if "https://www.google.com/" in fname and "google" not in fname:
                fnamelist = fname.split("/ar")
                fnamelist = fnamelist[0].split("/?trk")
                fnamelist = fnamelist[0].split("?trk")
                fnamelist = fnamelist[0].split("/%")
                set1.add((fnamelist[0]))

        matched = set1

        # Output these sets into new excel sheet
        output = Workbook()
        outSheet = output.active

        for item in matched:
            # outSheet.append((item[0], item[1], "matched"))
            # outSheet.append((item, "extracted"))
            outSheet.append((item, ""))

        # saves file in directory
        currdir = os.getcwd()
        dir_list = os.listdir(currdir)
        OutputDir = "outputdir"
        if OutputDir not in dir_list:
            os.mkdir(("outputdir"))
        outputdest = f"{currdir}/{OutputDir}"
        outputtitle = f1filename.replace("xlsx", "") + "updated"
        output.save(f"{outputdest}/{outputtitle}.xlsx")

        print("Analyzed and exported")


C = Compare()


def OpenFile() -> object:
    file1 = input("Enter file path: ")
    print("here", file1)
    f1 = C.setFile1(file1)

    # Gets name of file to append and sends it to C.rocess() to append to output
    filelocation = f1
    filelocationsplit = filelocation.split("/")
    f1filename = ""
    for items in filelocationsplit:
        if "xlsx" in items:
            f1filename = items
            print(items)

    if f1 is not None:
        try:
            C.process(f1filename)
        except:
            InvalidFileException

OpenFile()