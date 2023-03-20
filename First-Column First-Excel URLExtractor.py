#!/usr/bin/python3
# program cleans up XLSX files with URL and outputs in separate file
# file is placed in outputdirectory, program creates directory if it doesn't already exist

import os
from openpyxl import *
from tkinter.filedialog import askopenfilename
from tkinter import *
from tkinter import messagebox
from openpyxl.utils.exceptions import InvalidFileException  # handles situation when no file is selected
import os

root = Tk()

root.geometry('604x360+100+100')
root.title("Excel Comparison")

print("current directory", os.getcwd())

f1 = ""
# f2 = ""


class Compare():
    def __init__(self):
        self.file1 = ""
        # self.file2 = ""

    def setFile1(self, file1):
        print("File 1 set")
        self.file1 = file1
        self.change_pic1()
        return self.file1

    # def setFile2(self, file2):
    #     print("File 2 set")
    #     self.file2 = file2
    #     return self.file2

    # def __str__(self):
    #     return str(self.file1 + '\t' + self.file2)

    def process(self, f1filename):
        # Opening up the files
        file1 = load_workbook(self.file1)
        sheet1 = file1.active
        # file2 = load_workbook(self.file2)
        # sheet2 = file2.active

        # Initialize sets for storing the (first name, last name) for the two files
        set1 = set()
        set2 = set()

        # Iterate through rows of each file and add to set
        for i in range(1, sheet1.max_row + 1):
            fname = sheet1.cell(row=i, column=1).value.lower()
            # lname = sheet1.cell(row=i, column=2).value.lower()
            if "https://www.google.com/" in fname and "google" not in fname: # enter initial part of URL here
                fnamelist = fname.split("/ar")
                fnamelist = fnamelist[0].split("/?trk")
                fnamelist = fnamelist[0].split("?trk")
                fnamelist = fnamelist[0].split("/%")
                set1.add((fnamelist[0]))

        # for i in range(1, sheet2.max_row + 1):
        #     fname = sheet2.cell(row=i, column=1).value.lower()
        #     lname = sheet2.cell(row=i, column=2).value.lower()

        #     set2.add((fname, lname))

        matched = set1
        # matched = set1 & set2  # set1.intersection(set2)
        # unmatched1 = set1.difference(set2)  # set1 - set2
        # unmatched2 = set2.difference(set1)  # set2 - set1

        # Output these sets into new excel sheet
        output = Workbook()
        outSheet = output.active

        for item in matched:
            # outSheet.append((item[0], item[1], "matched"))
            # outSheet.append((item, "extracted"))
            outSheet.append((item, ""))

        # for item in unmatched1:
        #     outSheet.append(
        #         (item[0], item[1], "unmatched", "Only appears in file 1"))

        # for item in unmatched2:
        #     outSheet.append(
        #         (item[0], item[1], "unmatched", "Only appears in file 2"))

        # saves file in directory
        currdir = os.getcwd()
        dir_list = os.listdir(currdir)
        OutputDir = "outputdir"
        if OutputDir not in dir_list:
            os.mkdir(("outputdir"))
        outputdest = f"{currdir}/{OutputDir}"
        outputtitle = f1filename.replace("xlsx", "") + "updated"
        output.save(f"{outputdest}/{outputtitle}.xlsx")

        print("Analyzed and exported")

        C.finalMessage()  # notifies user process is completed

    def change_pic1(self):
        photo1 = PhotoImage(file=r'images/thumbnail_file_clicked.png')
        compose_button.configure(image=photo1)
        compose_button.photo = photo1
        print("updatedbutton1")

    # def change_pic2(self):
    #     photo1 = PhotoImage(file=r'thumbnail_file_clicked.png')
    #     compose_button2.configure(image=photo1)
    #     compose_button2.photo = photo1
    #     print("updatedbutton2")

    def finalMessage(self):
        # C.change_pic2()
        root.update()  # refreshes UI to update checked box thumbnail
        # time.sleep(2)
        messagebox.showinfo("", "Analyzed and exported")


C = Compare()


def OpenFile() -> object:
    file1 = askopenfilename(initialdir="C:/Users/Grant/Documents/Text/",
                            filetypes=(("All Files", "*.*"), ("All Files", "*.*")), title="Select a file (modded).")
    print("here", file1)
    f1 = C.setFile1(file1)

    # Gets name of file to append and sends it to C.rocess() to append to output
    filelocation = f1
    filelocationsplit = filelocation.split("/")
    f1filename = ""
    for items in filelocationsplit:
        if "xlsx" in items:
            f1filename = items
            print(items)

    if f1 is not None:
        try:
            C.process(f1filename)
        except:
            InvalidFileException


frame3 = Frame(root, width=200, height=150, background="white")
frame3.grid(row=0, column=1, rowspan=1, columnspan=50, sticky='w')

# def OpenFile2() -> object:
#     file2 = askopenfilename(initialdir="C:/Users/Grant/Documents/Text/",
#                             filetypes=(("All Files", "*.*"), ("All Files", "*.*")), title="Select a file (modded).")

#     f2 = C.setFile2(file2)
#     print("f2", f2)
#     if f2 is not None:
#         C.process()


prof_img = PhotoImage(
    file=r'images/Background.001.png')
file1image1 = PhotoImage(file=r'images/thumbnail_file.png')
file1image2 = PhotoImage(file=r'images/thumbnail_file.png')

lbl1 = Label(frame3, image=prof_img, compound=TOP)
lbl1.grid(rowspan=10, columnspan=40, column=0, row=0)

compose_button = Button(frame3, text="Select File 1",
                        image=file1image1, command=OpenFile)
compose_button.grid(column=17, row=5)

# compose_button2 = Button(frame3, text="Select File 2",
#                          image=file1image2, command=OpenFile2)
# compose_button2.grid(column=27, row=5)

root.mainloop()
